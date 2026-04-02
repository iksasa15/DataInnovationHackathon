"""
تحميل جدول قواعد الأعمال من LFS_Business_Rules.xlsx (جذر المستودع).
يُستخدم كمرجع لرسائل القواعد (EN) وأنواعها مع طبقة التحقق البرمجي في lfs_business_rules.py.
"""

from __future__ import annotations

import os
from functools import lru_cache
from pathlib import Path
from typing import Any

_REPO_ROOT = Path(__file__).resolve().parents[1]


def business_rules_xlsx_path() -> Path:
    """مسار الملف: متغير البيئة ثم جذر المستودع."""
    env = (os.getenv("LFS_BUSINESS_RULES_XLSX") or "").strip()
    if env:
        p = Path(env).expanduser()
        if p.is_file():
            return p.resolve()
    return (_REPO_ROOT / "LFS_Business_Rules.xlsx").resolve()


def _read_rows_openpyxl(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl مطلوب لقراءة LFS_Business_Rules.xlsx — ثبّت: pip install openpyxl"
        ) from e

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    def col(*candidates: str) -> int | None:
        for name in candidates:
            if name in header:
                return header.index(name)
        return None

    i_num = col("Error rule number")
    i_type = col("Error type")
    i_action = col("Action")
    i_summary = col("Rule summary (breakdown shown in hidden columns)")
    i_msg_en = col("Error message (EN)")
    if i_num is None:
        return []

    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if row is None:
            continue
        raw = row[i_num] if i_num < len(row) else None
        if raw is None or str(raw).strip() == "":
            continue
        try:
            rule_id = int(float(raw))
        except (TypeError, ValueError):
            continue

        def cell(idx: int | None) -> str:
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            if v is None:
                return ""
            return str(v).strip()

        out.append(
            {
                "rule_id": rule_id,
                "error_type": cell(i_type),
                "action": cell(i_action),
                "rule_summary": cell(i_summary),
                "message_en": cell(i_msg_en),
            }
        )
    return out


@lru_cache(maxsize=1)
def load_business_rules_table() -> tuple[dict[str, Any], ...]:
    """جميع الصفوف كـ tuples من dicts (قابل للتخزين المؤقت)."""
    path = business_rules_xlsx_path()
    if not path.is_file():
        return ()
    rows = _read_rows_openpyxl(path)
    return tuple(rows)


def invalidate_business_rules_cache() -> None:
    load_business_rules_table.cache_clear()
    rules_by_id.cache_clear()


@lru_cache(maxsize=1)
def rules_by_id() -> dict[int, dict[str, Any]]:
    return {r["rule_id"]: dict(r) for r in load_business_rules_table()}


def get_rule_record(rule_id: int) -> dict[str, Any] | None:
    return rules_by_id().get(int(rule_id))


def get_rule_message_en(rule_id: int) -> str:
    r = get_rule_record(rule_id)
    if not r:
        return ""
    return str(r.get("message_en") or "").strip()


def list_rules_for_api(limit: int = 500) -> list[dict[str, Any]]:
    """قائمة مبسّطة لواجهة REST والاختبارات."""
    out: list[dict[str, Any]] = []
    for r in load_business_rules_table():
        if len(out) >= limit:
            break
        out.append(
            {
                "rule_id": r["rule_id"],
                "error_type": r.get("error_type") or "",
                "action": (r.get("action") or "")[:200],
                "rule_summary": (r.get("rule_summary") or "")[:400],
                "message_en": (r.get("message_en") or "")[:800],
            }
        )
    return out


def business_rules_source_info() -> dict[str, Any]:
    path = business_rules_xlsx_path()
    return {
        "path": str(path),
        "exists": path.is_file(),
        "rule_count": len(load_business_rules_table()),
    }
